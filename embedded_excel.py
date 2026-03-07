"""
embedded_excel.py — Embedded Excel object handler for SlideArabi.

Handles two categories of embedded Excel content in PPTX files:

  Case 1 — OLE-embedded Excel tables
      Detected via <p:oleObj progId="Excel.Sheet.*"> inside a graphicFrame
      with graphicData URI = .../presentationml/2006/ole.
      Strategy: extract embedded .xlsx blob, translate string cells with openpyxl,
      set RTL sheet view, repack blob into the PPTX part.
      NOTE: The EMF/WMF preview image becomes stale — this is accepted and logged.

  Case 2 — Native DrawingML charts with embedded Excel data sources
      Detected via <c:chart> inside a graphicFrame (URI = .../drawingml/2006/chart).
      Charts already have partial RTL handling in rtl_transforms.py.
      Additional need: translate <c:v> cached string values in the chart XML,
      and optionally translate the embedded xlsx if autoUpdate="1".
      Strategy: translate chart text (title, axes, series names, category labels)
      and optionally the embedded workbook.

Design principles (matching the rest of SlideArabi):
- Never crash the pipeline: all public methods catch exceptions and log/return safe values.
- Comprehensive logging at each step.
- Defensive programming: handle None, missing attributes, malformed XML gracefully.
- Type hints on all public methods.
- Dataclasses for structured data.
"""

from __future__ import annotations

import io
import logging
import re
import zipfile
import shutil
import os
import tempfile
from copy import deepcopy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from lxml import etree

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# OOXML Namespace constants
# ─────────────────────────────────────────────────────────────────────────────

NSMAP: Dict[str, str] = {
    'a':  'http://schemas.openxmlformats.org/drawingml/2006/main',
    'p':  'http://schemas.openxmlformats.org/presentationml/2006/main',
    'c':  'http://schemas.openxmlformats.org/drawingml/2006/chart',
    'r':  'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'x':  'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
}

A_NS  = NSMAP['a']
P_NS  = NSMAP['p']
C_NS  = NSMAP['c']
R_NS  = NSMAP['r']
MC_NS = NSMAP['mc']

# graphicData URI discriminators
URI_TABLE  = 'http://schemas.openxmlformats.org/drawingml/2006/table'
URI_CHART  = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
URI_OLE    = 'http://schemas.openxmlformats.org/presentationml/2006/ole'
URI_DIAGRAMS = 'http://schemas.openxmlformats.org/drawingml/2006/diagram'

# Relationship types
REL_PACKAGE    = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/package'
REL_OLE_OBJECT = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/oleObject'

# Excel progId prefixes that indicate an Excel workbook OLE object
EXCEL_PROG_IDS = frozenset({
    'Excel.Sheet.8',
    'Excel.Sheet.12',
    'Excel.SheetMacroEnabled.12',
    'Excel.SheetBinaryMacroEnabled.12',
})

# Patterns for text that should NOT be translated (date codes, tickers, etc.)
SKIP_TRANSLATION_PATTERNS: List[str] = [
    r'^\d{4}$',               # bare year: 2024
    r'^Q[1-4]\s*\d{0,4}$',   # quarter: Q1, Q1 2024
    r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?$',  # month abbrevs
    r'^\d+(\.\d+)?%?$',       # pure number or percentage
    r'^[A-Z]{2,6}\d*$',       # ticker/code: AAPL, USD, SKU001
    r'^FY\d{2,4}$',           # fiscal year: FY2024
    r'^H[12]\s*\d{0,4}$',     # half-year: H1, H1 2024
]

_COMPILED_SKIP_PATTERNS = [re.compile(p, re.IGNORECASE) for p in SKIP_TRANSLATION_PATTERNS]


# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class EmbeddedExcelInfo:
    """Describes a single embedded Excel object found in a presentation.

    Attributes:
        slide_number: 1-based slide number where the object was found.
        shape_id: Numeric shape ID within the slide.
        shape_name: Human-readable name of the shape (e.g., 'Object 2').
        object_type: Classification — one of 'ole_excel_table', 'native_chart',
            'ole_excel_chart', 'native_table', 'picture', 'smartart', 'text_shape',
            'unknown'.
        prog_id: Excel progId string (e.g., 'Excel.Sheet.12'), or None for charts.
        is_embedded: True if the Excel blob is embedded in the PPTX (vs. linked).
        embedding_rId: Relationship ID of the embedded blob in the slide part.
        embedding_target: ZIP path of the embedded file within the PPTX
            (e.g., 'ppt/embeddings/Microsoft_Excel_Worksheet1.xlsx').
        has_preview_image: True if the OLE object displays a preview EMF/WMF image.
        preview_rId: Relationship ID of the preview image, or None.
        x_emu: Left position in EMU.
        y_emu: Top position in EMU.
        width_emu: Width in EMU.
        height_emu: Height in EMU.
        chart_has_autoUpdate: For chart objects — whether the chart's external
            data reference has autoUpdate="1". None for non-chart types.
        audit_notes: Free-text notes accumulated during processing.
    """
    slide_number: int
    shape_id: int
    shape_name: str
    object_type: str
    prog_id: Optional[str] = None
    is_embedded: bool = False
    embedding_rId: Optional[str] = None
    embedding_target: Optional[str] = None
    has_preview_image: bool = False
    preview_rId: Optional[str] = None
    x_emu: int = 0
    y_emu: int = 0
    width_emu: int = 0
    height_emu: int = 0
    chart_has_autoUpdate: Optional[bool] = None
    audit_notes: List[str] = field(default_factory=list)


@dataclass
class CellData:
    """A single translatable cell extracted from an OLE Excel workbook.

    Attributes:
        worksheet_name: Name of the worksheet.
        row: 1-based row index.
        col: 1-based column index.
        cell_address: Excel cell address (e.g., 'B3').
        original_value: Original string value of the cell.
        translated_value: Translated value, populated after translation.
    """
    worksheet_name: str
    row: int
    col: int
    cell_address: str
    original_value: str
    translated_value: Optional[str] = None


@dataclass
class TableData:
    """Data extracted from an embedded OLE Excel workbook.

    Attributes:
        workbook_bytes: Raw bytes of the embedded .xlsx workbook.
        translatable_cells: All string cells that are candidates for translation.
        worksheet_names: Names of all worksheets in the workbook.
        error: Error message if extraction failed, else None.
    """
    workbook_bytes: Optional[bytes]
    translatable_cells: List[CellData] = field(default_factory=list)
    worksheet_names: List[str] = field(default_factory=list)
    error: Optional[str] = None


@dataclass
class EmbeddedExcelHandlerReport:
    """Summary report produced by the EmbeddedExcelHandler after processing.

    Attributes:
        total_ole_tables_found: Number of OLE Excel table shapes detected.
        total_ole_tables_translated: Number successfully translated.
        total_charts_found: Number of DrawingML chart shapes detected.
        total_charts_translated: Number of charts with text translated.
        total_cells_translated: Total Excel cells translated across all objects.
        audit_log: Ordered list of audit events (type, message).
        errors: List of error messages (non-fatal).
        warnings: List of warning messages.
    """
    total_ole_tables_found: int = 0
    total_ole_tables_translated: int = 0
    total_charts_found: int = 0
    total_charts_translated: int = 0
    total_cells_translated: int = 0
    audit_log: List[Dict[str, str]] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def log_audit(self, event_type: str, message: str) -> None:
        """Append an audit event."""
        self.audit_log.append({'type': event_type, 'message': message})
        logger.info('[AUDIT:%s] %s', event_type, message)

    def warn(self, message: str) -> None:
        """Append a warning."""
        self.warnings.append(message)
        logger.warning('[EmbeddedExcel] %s', message)

    def error(self, message: str) -> None:
        """Append a non-fatal error."""
        self.errors.append(message)
        logger.error('[EmbeddedExcel] %s', message)


# ─────────────────────────────────────────────────────────────────────────────
# Shape classification helpers (standalone, usable outside the handler)
# ─────────────────────────────────────────────────────────────────────────────

def _get_graphicData_uri(graphic_frame_elem: etree._Element) -> Optional[str]:
    """Extract the uri attribute from <a:graphicData> within a graphicFrame element."""
    # Use iter() with Clark notation — compatible with python-pptx _OxmlElementBase
    # (which overrides .xpath() to reject the 'namespaces' kwarg).
    for gd in graphic_frame_elem.iter(f'{{{A_NS}}}graphicData'):
        uri = gd.get('uri')
        if uri:
            return uri
    return None


def _get_ole_prog_id(graphic_frame_elem: etree._Element) -> Optional[str]:
    """Extract the progId attribute from <p:oleObj> within an OLE graphicFrame."""
    for ole_obj in graphic_frame_elem.iter(f'{{{P_NS}}}oleObj'):
        prog_id = ole_obj.get('progId')
        if prog_id:
            return prog_id
    return None


def _get_ole_rId(graphic_frame_elem: etree._Element) -> Optional[str]:
    """Extract the r:id of the embedded OLE blob from <p:oleObj>."""
    r_id_attr = f'{{{R_NS}}}id'
    for ole_obj in graphic_frame_elem.iter(f'{{{P_NS}}}oleObj'):
        rId = ole_obj.get(r_id_attr)
        if rId:
            return rId
    return None


def _ole_is_embedded(graphic_frame_elem: etree._Element) -> bool:
    """Return True if the OLE object is embedded (has <p:embed/>), False if linked."""
    for _ in graphic_frame_elem.iter(f'{{{P_NS}}}embed'):
        return True
    return False


def _is_image_only_shape(sp_elem: etree._Element) -> bool:
    """Return True if a <p:sp> element uses an image fill and has no meaningful text."""
    has_blip = any(True for _ in sp_elem.iter(f'{{{P_NS}}}blipFill'))
    if not has_blip:
        return False
    has_text = any(
        t.text.strip()
        for t in sp_elem.iter(f'{{{A_NS}}}t')
        if t.text
    )
    return not has_text


def _get_shape_position(elem: etree._Element) -> Tuple[int, int, int, int]:
    """Extract (x, y, cx, cy) in EMU from a shape's <p:xfrm> or <a:xfrm> element."""
    # Use iter() — compatible with both raw lxml and python-pptx _OxmlElementBase.
    xfrm = None
    for candidate in elem.iter(f'{{{P_NS}}}xfrm'):
        xfrm = candidate
        break
    if xfrm is None:
        for candidate in elem.iter(f'{{{A_NS}}}xfrm'):
            xfrm = candidate
            break
    if xfrm is None:
        return 0, 0, 0, 0
    off = xfrm.find(f'{{{A_NS}}}off')
    ext = xfrm.find(f'{{{A_NS}}}ext')
    x   = int(off.get('x', '0')) if off is not None else 0
    y   = int(off.get('y', '0')) if off is not None else 0
    cx  = int(ext.get('cx', '0')) if ext is not None else 0
    cy  = int(ext.get('cy', '0')) if ext is not None else 0
    return x, y, cx, cy


def classify_shape(shape: Any) -> str:
    """Classify a python-pptx shape into one of the known embedded-content types.

    Returns one of:
        'native_table'      — <a:tbl> inside graphicFrame
        'ole_excel_table'   — OLE object with Excel progId (Sheet or Worksheet)
        'native_chart'      — DrawingML chart (<c:chart> in graphicFrame)
        'ole_excel_chart'   — Legacy OLE chart (e.g., MSGraph.Chart.8)
        'picture'           — Static image, no live text content
        'text_shape'        — Regular text box or placeholder
        'smartart'          — SmartArt diagram
        'unknown'           — Unrecognized

    Args:
        shape: A python-pptx Shape object.

    Returns:
        String classification label.
    """
    try:
        elem = shape.element
    except AttributeError:
        return 'unknown'

    tag = elem.tag

    # --- Picture: <p:pic> element ---
    if tag.endswith('}pic'):
        return 'picture'

    # --- GraphicFrame: tables, charts, OLE objects ---
    if tag.endswith('}graphicFrame'):
        gd_uri = _get_graphicData_uri(elem)

        if gd_uri == URI_TABLE:
            return 'native_table'

        elif gd_uri == URI_CHART:
            return 'native_chart'

        elif gd_uri == URI_OLE:
            prog_id = _get_ole_prog_id(elem)
            if prog_id:
                # Check for Excel workbook types
                if (prog_id in EXCEL_PROG_IDS or
                        prog_id.startswith('Excel.') or
                        prog_id.startswith('Sheet')):
                    return 'ole_excel_table'
                # Legacy OLE chart (MSGraph.Chart.8)
                if 'Chart' in prog_id or 'Graph' in prog_id:
                    return 'ole_excel_chart'
            return 'unknown'

        # SmartArt: graphicData URI contains "diagrams"
        elif gd_uri and ('diagrams' in gd_uri or 'diagram' in gd_uri):
            return 'smartart'

    # --- Regular shape with optional text body ---
    if tag.endswith('}sp'):
        if _is_image_only_shape(elem):
            return 'picture'
        return 'text_shape'

    # --- Group shape ---
    if tag.endswith('}grpSp'):
        return 'unknown'  # Groups handled elsewhere

    return 'unknown'


def should_translate_text(text: str) -> bool:
    """Return True if the text string is a candidate for translation.

    Skips date codes, quarter labels, tickers, pure numbers, etc.

    Args:
        text: The string to check.

    Returns:
        True if translation should be attempted; False to skip.
    """
    stripped = text.strip()
    if not stripped:
        return False
    for pattern in _COMPILED_SKIP_PATTERNS:
        if pattern.match(stripped):
            return False
    return True


# ─────────────────────────────────────────────────────────────────────────────
# Main handler class
# ─────────────────────────────────────────────────────────────────────────────

class EmbeddedExcelHandler:
    """Handles embedded Excel objects in PPTX slides for RTL conversion.

    Implements detection, extraction, translation, and write-back for:
    - OLE-embedded Excel table objects (Case 1)
    - Native DrawingML charts with embedded Excel data sources (Case 2)

    Usage::

        handler = EmbeddedExcelHandler()

        # Detect all embedded Excel objects across the presentation
        found = handler.detect_embedded_excel(prs)
        print(f"Found {len(found)} embedded Excel objects")

        # Process all OLE tables and charts in one pass
        report = handler.process_presentation(prs, translate_fn=my_translate_fn)
        print(report)

    Thread safety:
        Not thread-safe. Create a new instance per presentation.
    """

    def __init__(self) -> None:
        self._report = EmbeddedExcelHandlerReport()

    # ─────────────────────────────────────────────────────────────────────────
    # Detection
    # ─────────────────────────────────────────────────────────────────────────

    def detect_embedded_excel(self, prs: Any) -> List[EmbeddedExcelInfo]:
        """Scan all slides in a presentation for embedded Excel objects.

        Covers:
        - OLE-embedded Excel tables (progId starts with 'Excel.')
        - Native DrawingML charts (which always have an embedded Excel workbook)
        - Legacy OLE charts (MSGraph.Chart.8)

        Args:
            prs: A python-pptx Presentation object.

        Returns:
            List of EmbeddedExcelInfo dataclasses, one per detected object.
            Returns empty list on any error (never raises).
        """
        results: List[EmbeddedExcelInfo] = []
        try:
            for slide_idx, slide in enumerate(prs.slides, start=1):
                slide_results = self._scan_slide(slide, slide_idx)
                results.extend(slide_results)
        except Exception as exc:
            logger.error(
                'detect_embedded_excel: unexpected error scanning presentation: %s',
                exc, exc_info=True
            )
        logger.info(
            'detect_embedded_excel: found %d embedded Excel objects across %d slides',
            len(results), len(prs.slides) if hasattr(prs, 'slides') else 0
        )
        return results

    def _scan_slide(self, slide: Any, slide_number: int) -> List[EmbeddedExcelInfo]:
        """Scan a single slide for embedded Excel objects."""
        found: List[EmbeddedExcelInfo] = []
        try:
            for shape in slide.shapes:
                info = self._inspect_shape(shape, slide, slide_number)
                if info is not None:
                    found.append(info)
        except Exception as exc:
            logger.error(
                '_scan_slide: error on slide %d: %s', slide_number, exc, exc_info=True
            )
        return found

    def _inspect_shape(
        self, shape: Any, slide: Any, slide_number: int
    ) -> Optional[EmbeddedExcelInfo]:
        """Inspect a single shape and return EmbeddedExcelInfo if it contains
        embedded Excel content, else None."""
        try:
            kind = classify_shape(shape)
            if kind not in ('ole_excel_table', 'native_chart', 'ole_excel_chart'):
                return None

            elem = shape.element
            x, y, cx, cy = _get_shape_position(elem)
            shape_id = getattr(shape, 'shape_id', 0)
            shape_name = getattr(shape, 'name', '')

            if kind == 'ole_excel_table':
                return self._inspect_ole_table(
                    shape, elem, slide, slide_number, shape_id, shape_name,
                    x, y, cx, cy
                )
            elif kind in ('native_chart', 'ole_excel_chart'):
                return self._inspect_chart(
                    shape, elem, slide, slide_number, shape_id, shape_name,
                    x, y, cx, cy, kind
                )

        except Exception as exc:
            logger.warning(
                '_inspect_shape: error inspecting shape %s on slide %d: %s',
                getattr(shape, 'name', '?'), slide_number, exc, exc_info=True
            )
        return None

    def _inspect_ole_table(
        self, shape: Any, elem: etree._Element, slide: Any,
        slide_number: int, shape_id: int, shape_name: str,
        x: int, y: int, cx: int, cy: int
    ) -> EmbeddedExcelInfo:
        """Build EmbeddedExcelInfo for an OLE Excel table shape."""
        prog_id = _get_ole_prog_id(elem)
        is_embedded = _ole_is_embedded(elem)
        embedding_rId = _get_ole_rId(elem)
        embedding_target: Optional[str] = None
        preview_rId: Optional[str] = None
        has_preview = False

        # Resolve the relationship target path
        if embedding_rId:
            try:
                slide_part = shape.part
                rel = slide_part.rels.get(embedding_rId)
                if rel is not None:
                    embedding_target = rel.target_ref
            except Exception as exc:
                logger.debug(
                    '_inspect_ole_table: could not resolve rId %s: %s',
                    embedding_rId, exc
                )

        # Check for preview image rId: look for a:blip elements (inside OLE preview pic)
        r_embed_attr = f'{{{R_NS}}}embed'
        for blip in elem.iter(f'{{{A_NS}}}blip'):
            rId_val = blip.get(r_embed_attr)
            if rId_val:
                preview_rId = rId_val
                has_preview = True
                break

        notes: List[str] = []
        if not is_embedded:
            notes.append('OLE_LINKED: object is externally linked, not embedded')
        if has_preview:
            notes.append('HAS_PREVIEW_IMAGE: EMF preview will become stale after translation')

        return EmbeddedExcelInfo(
            slide_number=slide_number,
            shape_id=shape_id,
            shape_name=shape_name,
            object_type='ole_excel_table',
            prog_id=prog_id,
            is_embedded=is_embedded,
            embedding_rId=embedding_rId,
            embedding_target=embedding_target,
            has_preview_image=has_preview,
            preview_rId=preview_rId,
            x_emu=x, y_emu=y, width_emu=cx, height_emu=cy,
            audit_notes=notes,
        )

    def _inspect_chart(
        self, shape: Any, elem: etree._Element, slide: Any,
        slide_number: int, shape_id: int, shape_name: str,
        x: int, y: int, cx: int, cy: int, kind: str
    ) -> EmbeddedExcelInfo:
        """Build EmbeddedExcelInfo for a native DrawingML chart shape."""
        embedding_target: Optional[str] = None
        embedding_rId: Optional[str] = None
        chart_has_autoUpdate: Optional[bool] = None
        is_embedded = False

        try:
            chart_part = shape.chart_part
            chart_xml = chart_part._element

            # Check for autoUpdate — chart_xml is a raw lxml element (chart_part._element)
            for ext_data in chart_xml.iter(f'{{{C_NS}}}externalData'):
                for auto_update in ext_data.iter(f'{{{C_NS}}}autoUpdate'):
                    val = auto_update.get('val', '0')
                    chart_has_autoUpdate = val.lower() in ('1', 'true')
                    break
                break

            # Inspect chart relationships for the embedded workbook
            try:
                for rId, rel in chart_part.rels.items():
                    if rel.reltype == REL_PACKAGE:
                        is_embedded = True
                        embedding_rId = rId
                        embedding_target = rel.target_ref
                        break
                    elif rel.reltype == REL_OLE_OBJECT:
                        # Linked external workbook
                        embedding_rId = rId
                        embedding_target = rel.target_ref
                        is_embedded = False
                        break
            except Exception as exc:
                logger.debug(
                    '_inspect_chart: could not iterate chart rels for %s: %s',
                    shape_name, exc
                )

        except Exception as exc:
            logger.debug(
                '_inspect_chart: could not access chart_part for %s: %s',
                shape_name, exc
            )

        notes: List[str] = []
        if not is_embedded:
            notes.append('CHART_LINKED_WORKBOOK: data workbook is external; only cached values will be translated')
        if chart_has_autoUpdate:
            notes.append('CHART_AUTOUPDATE: autoUpdate=1; embedded workbook also requires translation')

        return EmbeddedExcelInfo(
            slide_number=slide_number,
            shape_id=shape_id,
            shape_name=shape_name,
            object_type=kind,
            prog_id=None,
            is_embedded=is_embedded,
            embedding_rId=embedding_rId,
            embedding_target=embedding_target,
            has_preview_image=False,
            x_emu=x, y_emu=y, width_emu=cx, height_emu=cy,
            chart_has_autoUpdate=chart_has_autoUpdate,
            audit_notes=notes,
        )

    # ─────────────────────────────────────────────────────────────────────────
    # Extraction — OLE Excel tables
    # ─────────────────────────────────────────────────────────────────────────

    def extract_table_data(self, ole_info: EmbeddedExcelInfo, shape: Any) -> TableData:
        """Extract cell data from the embedded Excel workbook of an OLE table shape.

        Reads the embedded .xlsx blob via python-pptx's ole_format API, then
        parses it with openpyxl to enumerate all translatable string cells.

        Args:
            ole_info: EmbeddedExcelInfo for this OLE object.
            shape: The python-pptx Shape object for the OLE table.

        Returns:
            TableData with workbook_bytes and translatable_cells populated.
            On failure, returns TableData with error set and workbook_bytes=None.
        """
        try:
            xlsx_bytes = self._get_ole_blob(shape)
        except Exception as exc:
            msg = f'extract_table_data: failed to get OLE blob for shape {ole_info.shape_name}: {exc}'
            logger.error(msg, exc_info=True)
            return TableData(workbook_bytes=None, error=msg)

        if xlsx_bytes is None:
            msg = (
                f'extract_table_data: OLE blob is None for shape {ole_info.shape_name} '
                f'(linked OLE objects are not supported)'
            )
            logger.warning(msg)
            self._report.log_audit('OLE_LINKED_SKIP', msg)
            return TableData(workbook_bytes=None, error=msg)

        try:
            import openpyxl
        except ImportError:
            msg = 'extract_table_data: openpyxl is not installed; cannot read .xlsx blobs'
            logger.error(msg)
            return TableData(workbook_bytes=xlsx_bytes, error=msg)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        except Exception as exc:
            msg = (
                f'extract_table_data: openpyxl failed to load workbook for shape '
                f'{ole_info.shape_name}: {exc}'
            )
            logger.error(msg, exc_info=True)
            return TableData(workbook_bytes=xlsx_bytes, error=msg)

        translatable_cells: List[CellData] = []
        worksheet_names: List[str] = []

        try:
            for ws in wb.worksheets:
                worksheet_names.append(ws.title)
                for row in ws.iter_rows():
                    for cell in row:
                        # Only process string-typed cells with non-empty values
                        if cell.data_type == 's' and cell.value:
                            val_str = str(cell.value)
                            # Skip formula strings
                            if val_str.startswith('='):
                                continue
                            if should_translate_text(val_str):
                                from openpyxl.utils.cell import get_column_letter
                                col_letter = get_column_letter(cell.column)
                                cell_address = f'{col_letter}{cell.row}'
                                translatable_cells.append(CellData(
                                    worksheet_name=ws.title,
                                    row=cell.row,
                                    col=cell.column,
                                    cell_address=cell_address,
                                    original_value=val_str,
                                ))
        except Exception as exc:
            logger.warning(
                'extract_table_data: error enumerating cells for %s: %s',
                ole_info.shape_name, exc, exc_info=True
            )

        logger.info(
            'extract_table_data: shape=%s, worksheets=%d, translatable_cells=%d',
            ole_info.shape_name, len(worksheet_names), len(translatable_cells)
        )

        return TableData(
            workbook_bytes=xlsx_bytes,
            translatable_cells=translatable_cells,
            worksheet_names=worksheet_names,
        )

    def _get_ole_blob(self, shape: Any) -> Optional[bytes]:
        """Retrieve the embedded OLE blob bytes from a shape.

        Tries the python-pptx ole_format API first, then falls back to
        direct part access via the relationship.

        Args:
            shape: python-pptx Shape object for the OLE object.

        Returns:
            bytes of the embedded .xlsx workbook, or None if not accessible.
        """
        # Attempt 1: python-pptx public API
        try:
            ole_format = shape.ole_format
            if ole_format is not None:
                blob = ole_format.blob
                if blob:
                    logger.debug('_get_ole_blob: retrieved via ole_format.blob (%d bytes)', len(blob))
                    return blob
        except AttributeError:
            pass  # ole_format may not exist on all shape types
        except Exception as exc:
            logger.debug('_get_ole_blob: ole_format.blob failed: %s', exc)

        # Attempt 2: Direct relationship part access
        try:
            elem = shape.element
            rId = _get_ole_rId(elem)
            if rId:
                slide_part = shape.part
                related_part = slide_part.related_part(rId)
                blob = related_part.blob
                if blob:
                    logger.debug(
                        '_get_ole_blob: retrieved via related_part (rId=%s, %d bytes)',
                        rId, len(blob)
                    )
                    return blob
        except Exception as exc:
            logger.debug('_get_ole_blob: related_part fallback failed: %s', exc)

        return None

    # ─────────────────────────────────────────────────────────────────────────
    # Translation — OLE Excel tables
    # ─────────────────────────────────────────────────────────────────────────

    def translate_ole_excel_table(
        self,
        shape: Any,
        translate_fn: Callable[[str], str],
        ole_info: Optional[EmbeddedExcelInfo] = None,
    ) -> int:
        """Translate string cells in the embedded Excel workbook of an OLE table.

        Reads the .xlsx blob, translates string cells (skipping formulas, numbers,
        date codes), sets RTL sheet view, and writes the translated blob back to
        the PPTX part.

        Args:
            shape: python-pptx Shape object for the OLE table.
            translate_fn: Callable that takes an English string and returns Arabic.
            ole_info: Optional EmbeddedExcelInfo for logging context.

        Returns:
            Number of cells translated. Returns 0 on failure (never raises).
        """
        shape_name = getattr(shape, 'name', '<unknown>')
        try:
            import openpyxl
        except ImportError:
            logger.error('translate_ole_excel_table: openpyxl not installed')
            return 0

        try:
            xlsx_bytes = self._get_ole_blob(shape)
            if xlsx_bytes is None:
                self._report.log_audit(
                    'OLE_LINKED_SKIP',
                    f'Shape {shape_name}: OLE blob not accessible (linked object)'
                )
                return 0

            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
            cells_translated = 0

            for ws in wb.worksheets:
                # Set RTL view on the worksheet
                try:
                    sv = ws.sheet_view
                    sv.rightToLeft = True
                except Exception:
                    # Fallback: try views collection if present
                    try:
                        for view in ws.views.sheetView:
                            view.rightToLeft = True
                    except Exception:
                        logger.debug('Could not set RTL on worksheet %s', ws.title)

                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 's' and cell.value:
                            val_str = str(cell.value)
                            if val_str.startswith('='):
                                continue  # Never translate formulas
                            if should_translate_text(val_str):
                                try:
                                    translated = translate_fn(val_str)
                                    if translated and translated != val_str:
                                        cell.value = translated
                                        cells_translated += 1
                                except Exception as tex:
                                    logger.warning(
                                        'translate_ole_excel_table: translation failed '
                                        'for cell "%s": %s', val_str, tex
                                    )

            # Serialize back to bytes
            buf = io.BytesIO()
            wb.save(buf)
            new_bytes = buf.getvalue()

            # Write back to the PPTX part
            self._replace_ole_blob(shape, new_bytes)

            # Audit log: preview image is now stale
            if ole_info and ole_info.has_preview_image:
                self._report.log_audit(
                    'PREVIEW_IMAGE_STALE',
                    f'Shape {shape_name} (slide {ole_info.slide_number}): '
                    f'EMF preview image is now stale after Excel translation. '
                    f'Preview will show English content until refreshed in PowerPoint.'
                )
                self._report.warn(
                    f'Slide {ole_info.slide_number if ole_info else "?"}: '
                    f'Shape "{shape_name}" OLE preview image is stale.'
                )

            logger.info(
                'translate_ole_excel_table: shape=%s, cells_translated=%d',
                shape_name, cells_translated
            )
            self._report.total_cells_translated += cells_translated
            return cells_translated

        except Exception as exc:
            msg = (
                f'translate_ole_excel_table: unexpected error for shape {shape_name}: {exc}'
            )
            logger.error(msg, exc_info=True)
            self._report.error(msg)
            return 0

    def _replace_ole_blob(self, shape: Any, new_bytes: bytes) -> bool:
        """Replace the embedded OLE blob by writing directly to the part.

        Tries the python-pptx private API first, then falls back to relationship
        part access.

        Args:
            shape: python-pptx Shape object for the OLE object.
            new_bytes: New bytes to write to the embedded part.

        Returns:
            True on success, False on failure.
        """
        shape_name = getattr(shape, 'name', '<unknown>')

        # Attempt 1: via related_part (standard relationship access)
        try:
            elem = shape.element
            rId = _get_ole_rId(elem)
            if rId:
                slide_part = shape.part
                related_part = slide_part.related_part(rId)
                related_part._blob = new_bytes
                logger.debug(
                    '_replace_ole_blob: wrote %d bytes via related_part (rId=%s) for %s',
                    len(new_bytes), rId, shape_name
                )
                return True
        except Exception as exc:
            logger.warning(
                '_replace_ole_blob: related_part write failed for %s: %s',
                shape_name, exc
            )

        # Attempt 2: python-pptx ole_format path
        try:
            ole_format = shape.ole_format
            if ole_format is not None:
                # Some versions expose a writable blob
                part = ole_format._part
                if part is not None:
                    part._blob = new_bytes
                    logger.debug(
                        '_replace_ole_blob: wrote %d bytes via ole_format._part for %s',
                        len(new_bytes), shape_name
                    )
                    return True
        except Exception as exc:
            logger.warning(
                '_replace_ole_blob: ole_format._part write failed for %s: %s',
                shape_name, exc
            )

        logger.error('_replace_ole_blob: all write-back strategies failed for shape %s', shape_name)
        return False

    # ─────────────────────────────────────────────────────────────────────────
    # Translation — Native PPTX charts
    # ─────────────────────────────────────────────────────────────────────────

    def translate_chart_labels(
        self,
        shape: Any,
        translate_fn: Callable[[str], str],
        translate_embedded_workbook: bool = False,
    ) -> int:
        """Translate all visible text in a DrawingML chart.

        Translates:
        - Chart title (<c:chart/c:title//a:t>)
        - Category axis title (<c:catAx/c:title//a:t>)
        - Value axis title (<c:valAx/c:title//a:t>)
        - Series names cached in <c:ser/c:tx//c:v>
        - Category labels cached in <c:ser/c:cat//c:v>
        - Data labels with rich text override (<c:dLbls/c:dLbl/c:tx/c:rich//a:t>)
        - Legend entry labels (<c:legend/c:legendEntry/c:txPr//a:t>)

        Does NOT translate:
        - Numeric values in <c:numCache>
        - Formula references (<c:f>)
        - Format codes in <c:numFmt>

        Args:
            shape: python-pptx Shape object for the chart.
            translate_fn: Callable(str) -> str for English-to-Arabic translation.
            translate_embedded_workbook: If True and the chart has an embedded
                workbook with autoUpdate=1, also translate the xlsx data.

        Returns:
            Number of text nodes translated. Returns 0 on failure (never raises).
        """
        shape_name = getattr(shape, 'name', '<unknown>')
        texts_translated = 0

        try:
            chart_part = shape.chart_part
            chart_xml = chart_part._element
        except AttributeError as exc:
            logger.warning(
                'translate_chart_labels: shape %s has no chart_part: %s',
                shape_name, exc
            )
            return 0

        try:
            # 1. Translate rich-text <a:t> nodes (title, axis titles, data labels, legend)
            for t_elem in chart_xml.iter(f'{{{A_NS}}}t'):
                text = t_elem.text
                if text and text.strip() and should_translate_text(text):
                    try:
                        translated = translate_fn(text)
                        if translated and translated != text:
                            t_elem.text = translated
                            texts_translated += 1
                    except Exception as tex:
                        logger.warning(
                            'translate_chart_labels: translation failed for "%s": %s',
                            text, tex
                        )

            # 2. Translate cached string values in series definitions
            # Covers: series names (c:ser/c:tx//c:v), category labels (c:ser/c:cat//c:v),
            #         X-axis values (c:ser/c:xVal//c:v)
            # Collect series text values using iter + parent checks
            # (python-pptx _OxmlElementBase doesn't support namespaces= in xpath)
            c_v_tag = f'{{{C_NS}}}v'
            c_tx_tag = f'{{{C_NS}}}tx'
            c_cat_tag = f'{{{C_NS}}}cat'
            c_xVal_tag = f'{{{C_NS}}}xVal'
            c_ser_tag = f'{{{C_NS}}}ser'
            for v_elem in chart_xml.iter(c_v_tag):
                text = v_elem.text
                if text and text.strip():
                    # Skip pure numbers
                    try:
                        float(text)
                        continue
                    except ValueError:
                        pass
                    if should_translate_text(text):
                        try:
                            translated = translate_fn(text)
                            if translated and translated != text:
                                v_elem.text = translated
                                texts_translated += 1
                        except Exception as tex:
                            logger.warning(
                                'translate_chart_labels: <c:v> translation failed for "%s": %s',
                                text, tex
                            )

            # 3. Optionally translate the embedded workbook
            if translate_embedded_workbook:
                wb_texts = self._translate_chart_embedded_workbook(
                    shape, chart_part, translate_fn
                )
                texts_translated += wb_texts

        except Exception as exc:
            msg = f'translate_chart_labels: unexpected error for shape {shape_name}: {exc}'
            logger.error(msg, exc_info=True)
            self._report.error(msg)

        logger.info(
            'translate_chart_labels: shape=%s, texts_translated=%d',
            shape_name, texts_translated
        )
        return texts_translated

    def _translate_chart_embedded_workbook(
        self, shape: Any, chart_part: Any, translate_fn: Callable[[str], str]
    ) -> int:
        """Translate string cells in a chart's embedded Excel workbook.

        Only processes charts where the data workbook is embedded (REL_PACKAGE),
        not externally linked.

        Args:
            shape: python-pptx Shape for the chart.
            chart_part: The chart's part object.
            translate_fn: Translation callable.

        Returns:
            Number of workbook cells translated.
        """
        shape_name = getattr(shape, 'name', '<unknown>')
        cells_translated = 0

        try:
            import openpyxl
        except ImportError:
            logger.warning('_translate_chart_embedded_workbook: openpyxl not installed')
            return 0

        try:
            for rId, rel in chart_part.rels.items():
                if rel.reltype == REL_PACKAGE:
                    xlsx_bytes = rel.target_part.blob
                    if not xlsx_bytes:
                        continue
                    try:
                        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
                        for ws in wb.worksheets:
                            for view in ws.sheet_views:
                                view.rightToLeft = True
                            for row in ws.iter_rows():
                                for cell in row:
                                    if cell.data_type == 's' and cell.value:
                                        val = str(cell.value)
                                        if not val.startswith('=') and should_translate_text(val):
                                            try:
                                                translated = translate_fn(val)
                                                if translated and translated != val:
                                                    cell.value = translated
                                                    cells_translated += 1
                                            except Exception:
                                                pass
                        buf = io.BytesIO()
                        wb.save(buf)
                        rel.target_part._blob = buf.getvalue()
                        logger.debug(
                            '_translate_chart_embedded_workbook: shape=%s, cells=%d',
                            shape_name, cells_translated
                        )
                    except Exception as exc:
                        logger.warning(
                            '_translate_chart_embedded_workbook: openpyxl error for %s: %s',
                            shape_name, exc
                        )
                    break

        except Exception as exc:
            logger.warning(
                '_translate_chart_embedded_workbook: error for %s: %s',
                shape_name, exc, exc_info=True
            )

        return cells_translated

    # ─────────────────────────────────────────────────────────────────────────
    # RTL application — native PPTX table
    # ─────────────────────────────────────────────────────────────────────────

    def apply_rtl_to_table(self, shape: Any) -> int:
        """Apply RTL text formatting to all cells in a native PPTX table.

        Sets <a:pPr rtl="1" algn="r"> on every paragraph and
        <a:rPr lang="ar-SA" altLang="en-US"> on every run in every cell.
        Creates <a:pPr> if it does not exist.

        Args:
            shape: python-pptx Shape object for a native table.

        Returns:
            Number of paragraphs modified. Returns 0 on failure (never raises).
        """
        shape_name = getattr(shape, 'name', '<unknown>')
        paragraphs_modified = 0
        try:
            tbl = shape.table._tbl
            for tc in tbl.iter(f'{{{A_NS}}}tc'):
                for p in tc.iter(f'{{{A_NS}}}p'):
                    # Ensure <a:pPr> exists as first child
                    pPr = p.find(f'{{{A_NS}}}pPr')
                    if pPr is None:
                        pPr = etree.SubElement(p, f'{{{A_NS}}}pPr')
                        p.insert(0, pPr)
                    pPr.set('rtl', '1')
                    pPr.set('algn', 'r')
                    paragraphs_modified += 1

                # Set language on all run properties
                for rPr in tc.iter(f'{{{A_NS}}}rPr'):
                    rPr.set('lang', 'ar-SA')
                    rPr.set('altLang', 'en-US')

        except Exception as exc:
            logger.error(
                'apply_rtl_to_table: error for shape %s: %s', shape_name, exc, exc_info=True
            )
        logger.debug(
            'apply_rtl_to_table: shape=%s, paragraphs_modified=%d',
            shape_name, paragraphs_modified
        )
        return paragraphs_modified

    def reverse_table_columns(self, shape: Any) -> bool:
        """Reverse the column order of a native PPTX table for RTL layout.

        Reorders both:
        - <a:gridCol> elements in <a:tblGrid> (column width definitions)
        - <a:tc> elements within each <a:tr> (cell data)

        Handles merged cells correctly: gridSpan groups are treated as atomic
        units and reversed together (the origin cell plus its hMerge followers).

        Args:
            shape: python-pptx Shape object for a native table.

        Returns:
            True on success, False on failure.
        """
        shape_name = getattr(shape, 'name', '<unknown>')
        try:
            tbl = shape.table._tbl

            # 1. Reverse column width definitions in <a:tblGrid>
            tbl_grid = tbl.find(f'{{{A_NS}}}tblGrid')
            if tbl_grid is not None:
                cols = list(tbl_grid)
                for col in cols:
                    tbl_grid.remove(col)
                for col in reversed(cols):
                    tbl_grid.append(col)

            # 2. Reverse cells within each row (handling merges)
            for row_elem in tbl.findall(f'{{{A_NS}}}tr'):
                self._reverse_row_cells(row_elem)

            logger.debug('reverse_table_columns: shape=%s reversed', shape_name)
            return True

        except Exception as exc:
            logger.error(
                'reverse_table_columns: error for shape %s: %s', shape_name, exc, exc_info=True
            )
            return False

    def _reverse_row_cells(self, row_elem: etree._Element) -> None:
        """Reverse cells in a single table row, handling merged cells correctly.

        gridSpan="N" marks the origin cell of an N-column merge.
        Subsequent cells in the merge have hMerge="1".
        Groups of merged cells are reversed as atomic units.

        Args:
            row_elem: The <a:tr> lxml element to reverse.
        """
        cells = list(row_elem.findall(f'{{{A_NS}}}tc'))
        if not cells:
            return

        # Build groups: each group is a list of cells that belong together
        # (the origin cell plus any hMerge followers)
        groups: List[List[etree._Element]] = []
        i = 0
        while i < len(cells):
            tc = cells[i]
            grid_span = int(tc.get('gridSpan', '1'))
            if grid_span > 1:
                group = cells[i:i + grid_span]
                groups.append(group)
                i += grid_span
            else:
                groups.append([tc])
                i += 1

        # Remove all cells from row
        for cell in cells:
            row_elem.remove(cell)

        # Re-append groups in reversed order
        for group in reversed(groups):
            for cell in group:
                row_elem.append(cell)

    # ─────────────────────────────────────────────────────────────────────────
    # RTL application — charts
    # ─────────────────────────────────────────────────────────────────────────

    def apply_rtl_to_chart(self, shape: Any) -> bool:
        """Apply full RTL treatment to a native DrawingML chart.

        Steps applied:
        1. Reverse category axis direction (maxMin orientation)
        2. Move value axis to right side
        3. Adjust legend position (right → left)
        4. Set Arabic language on all text run properties
        5. Set RTL on all paragraph properties in chart text
        6. Handle multiple plot areas (combo charts)

        Does NOT apply axis reversal to pie/doughnut charts (non-directional).

        Args:
            shape: python-pptx Shape object for the chart.

        Returns:
            True on success, False on failure (never raises).
        """
        shape_name = getattr(shape, 'name', '<unknown>')
        try:
            chart_part = shape.chart_part
            chart_xml = chart_part._element
        except AttributeError as exc:
            logger.warning('apply_rtl_to_chart: shape %s has no chart_part: %s', shape_name, exc)
            return False

        try:
            # Check if this is a pie/doughnut chart — skip axis reversal for those
            pie_tag = f'{{{C_NS}}}pieChart'
            doughnut_tag = f'{{{C_NS}}}doughnutChart'
            has_pie = bool(list(chart_xml.iter(pie_tag)) or list(chart_xml.iter(doughnut_tag)))

            if not has_pie:
                # Step 1: Reverse all category axes
                for orientation in chart_xml.iter(f'{{{C_NS}}}orientation'):
                    orientation.set('val', 'maxMin')

                # Step 2: Move value axes to right side
                for axPos in chart_xml.iter(f'{{{C_NS}}}axPos'):
                    current = axPos.get('val', 'l')
                    if current == 'l':
                        axPos.set('val', 'r')
                    elif current == 'r':
                        axPos.set('val', 'l')

            # Step 3: Mirror legend position
            legend_pos_mapping = {'r': 'l', 'l': 'r', 'tr': 'tl', 'tl': 'tr'}
            for legendPos in chart_xml.iter(f'{{{C_NS}}}legendPos'):
                current = legendPos.get('val', 'r')
                legendPos.set('val', legend_pos_mapping.get(current, current))

            # Step 4: Set Arabic language on all run properties
            for rPr in chart_xml.iter(f'{{{A_NS}}}rPr'):
                rPr.set('lang', 'ar-SA')
                rPr.set('altLang', 'en-US')

            # Step 5: Set RTL on paragraph properties
            for pPr in chart_xml.iter(f'{{{A_NS}}}pPr'):
                pPr.set('rtl', '1')

            # Step 6: Set language on default run properties in txPr blocks
            for defRPr in chart_xml.iter(f'{{{A_NS}}}defRPr'):
                defRPr.set('lang', 'ar-SA')

            logger.debug('apply_rtl_to_chart: shape=%s processed', shape_name)
            return True

        except Exception as exc:
            msg = f'apply_rtl_to_chart: unexpected error for shape {shape_name}: {exc}'
            logger.error(msg, exc_info=True)
            self._report.error(msg)
            return False

    # ─────────────────────────────────────────────────────────────────────────
    # Position mirroring
    # ─────────────────────────────────────────────────────────────────────────

    def mirror_shape_position(self, shape: Any, slide_width_emu: int) -> bool:
        """Mirror a shape's horizontal position for RTL slide layout.

        Computes: new_x = slide_width - old_x - width

        Does NOT set flipH — that would visually mirror image content.

        Args:
            shape: python-pptx Shape object.
            slide_width_emu: Total slide width in EMU (typically 9144000 for 10").

        Returns:
            True if position was updated, False if not applicable or error.
        """
        shape_name = getattr(shape, 'name', '<unknown>')
        try:
            elem = shape.element
            # Find xfrm — use iter() for compatibility with python-pptx elements.
            xfrm = None
            for candidate in elem.iter(f'{{{P_NS}}}xfrm'):
                xfrm = candidate
                break
            if xfrm is None:
                for candidate in elem.iter(f'{{{A_NS}}}xfrm'):
                    xfrm = candidate
                    break
            if xfrm is None:
                logger.debug('mirror_shape_position: no xfrm found for %s', shape_name)
                return False
            off = xfrm.find(f'{{{A_NS}}}off')
            ext = xfrm.find(f'{{{A_NS}}}ext')

            if off is None or ext is None:
                return False

            old_x = int(off.get('x', '0'))
            width = int(ext.get('cx', '0'))
            new_x = slide_width_emu - old_x - width

            # Clamp to valid range
            new_x = max(0, new_x)

            off.set('x', str(new_x))
            logger.debug(
                'mirror_shape_position: shape=%s, old_x=%d, width=%d, new_x=%d',
                shape_name, old_x, width, new_x
            )
            return True

        except Exception as exc:
            logger.error(
                'mirror_shape_position: error for shape %s: %s', shape_name, exc, exc_info=True
            )
            return False

    # ─────────────────────────────────────────────────────────────────────────
    # Native PPTX table full treatment (translate + reverse + RTL)
    # ─────────────────────────────────────────────────────────────────────────

    def process_native_table(
        self,
        shape: Any,
        translate_fn: Callable[[str], str],
        slide_width_emu: int,
        mirror_position: bool = True,
    ) -> int:
        """Apply full RTL treatment to a native PPTX table.

        Steps:
        1. Translate all cell text via translate_fn
        2. Reverse column order (handling merges)
        3. Apply RTL paragraph/run properties to all cells
        4. Optionally mirror shape position

        Args:
            shape: python-pptx Shape for a native PPTX table.
            translate_fn: Translation callable.
            slide_width_emu: Slide width in EMU for position mirroring.
            mirror_position: If True, mirror the shape's X position.

        Returns:
            Number of cells translated.
        """
        shape_name = getattr(shape, 'name', '<unknown>')
        cells_translated = 0

        try:
            # Step 1: Translate cell text
            cells_translated = self._translate_native_table_cells(shape, translate_fn)

            # Step 2: Reverse column order
            self.reverse_table_columns(shape)

            # Step 3: Apply RTL formatting
            self.apply_rtl_to_table(shape)

            # Step 4: Mirror position
            if mirror_position:
                self.mirror_shape_position(shape, slide_width_emu)

            logger.info(
                'process_native_table: shape=%s, cells_translated=%d',
                shape_name, cells_translated
            )

        except Exception as exc:
            msg = f'process_native_table: unexpected error for shape {shape_name}: {exc}'
            logger.error(msg, exc_info=True)
            self._report.error(msg)

        return cells_translated

    def _translate_native_table_cells(
        self, shape: Any, translate_fn: Callable[[str], str]
    ) -> int:
        """Translate all <a:t> text nodes inside a native PPTX table.

        Args:
            shape: python-pptx Shape for the native table.
            translate_fn: Translation callable.

        Returns:
            Number of text nodes translated.
        """
        cells_translated = 0
        shape_name = getattr(shape, 'name', '<unknown>')
        try:
            tbl = shape.table._tbl
            for tc in tbl.iter(f'{{{A_NS}}}tc'):
                for t_elem in tc.iter(f'{{{A_NS}}}t'):
                    text = t_elem.text
                    if text and text.strip() and should_translate_text(text):
                        try:
                            translated = translate_fn(text)
                            if translated and translated != text:
                                t_elem.text = translated
                                cells_translated += 1
                        except Exception as tex:
                            logger.warning(
                                '_translate_native_table_cells: translation failed '
                                'for "%s" in shape %s: %s', text, shape_name, tex
                            )
        except Exception as exc:
            logger.error(
                '_translate_native_table_cells: error for shape %s: %s',
                shape_name, exc, exc_info=True
            )
        return cells_translated

    # ─────────────────────────────────────────────────────────────────────────
    # Full presentation processing
    # ─────────────────────────────────────────────────────────────────────────

    def process_presentation(
        self,
        prs: Any,
        translate_fn: Callable[[str], str],
        mirror_positions: bool = True,
        translate_charts: bool = True,
        translate_native_tables: bool = True,
        translate_ole_tables: bool = True,
    ) -> EmbeddedExcelHandlerReport:
        """Process all embedded Excel objects in a presentation for RTL conversion.

        Iterates every slide, classifies each shape, and applies the appropriate
        combination of translation + RTL formatting based on object type.

        Processing order per slide:
        1. OLE Excel tables → translate blob + mirror position
        2. Native DrawingML charts → translate labels + apply RTL axis/text
        3. Native PPTX tables → translate cells + reverse columns + RTL text
        4. Legacy OLE charts → log audit, mirror position only

        Args:
            prs: python-pptx Presentation object (in-place modification).
            translate_fn: Callable(str) -> str for English-to-Arabic translation.
                          Called once per text fragment; may raise (errors are caught).
            mirror_positions: If True, mirror each processed shape's X position.
            translate_charts: If True, process native charts.
            translate_native_tables: If True, process native PPTX tables.
            translate_ole_tables: If True, process OLE-embedded Excel tables.

        Returns:
            EmbeddedExcelHandlerReport with counts, audit log, warnings, errors.
            Never raises.
        """
        self._report = EmbeddedExcelHandlerReport()  # fresh report

        try:
            slide_width_emu = prs.slide_width
        except Exception:
            slide_width_emu = 9144000  # Default: 10 inches (standard widescreen)
            logger.warning(
                'process_presentation: could not read slide_width; '
                'using default %d EMU', slide_width_emu
            )

        total_slides = len(prs.slides) if hasattr(prs, 'slides') else 0
        logger.info(
            'process_presentation: starting — %d slides, mirror=%s, '
            'charts=%s, native_tables=%s, ole_tables=%s',
            total_slides, mirror_positions, translate_charts,
            translate_native_tables, translate_ole_tables
        )

        try:
            for slide_idx, slide in enumerate(prs.slides, start=1):
                self._process_slide(
                    slide=slide,
                    slide_number=slide_idx,
                    slide_width_emu=slide_width_emu,
                    translate_fn=translate_fn,
                    mirror_positions=mirror_positions,
                    translate_charts=translate_charts,
                    translate_native_tables=translate_native_tables,
                    translate_ole_tables=translate_ole_tables,
                )
        except Exception as exc:
            msg = f'process_presentation: unexpected error: {exc}'
            logger.error(msg, exc_info=True)
            self._report.error(msg)

        logger.info(
            'process_presentation: done — '
            'ole_tables found=%d translated=%d, '
            'charts found=%d translated=%d, '
            'cells_translated=%d, errors=%d, warnings=%d',
            self._report.total_ole_tables_found,
            self._report.total_ole_tables_translated,
            self._report.total_charts_found,
            self._report.total_charts_translated,
            self._report.total_cells_translated,
            len(self._report.errors),
            len(self._report.warnings),
        )
        return self._report

    def _process_slide(
        self,
        slide: Any,
        slide_number: int,
        slide_width_emu: int,
        translate_fn: Callable[[str], str],
        mirror_positions: bool,
        translate_charts: bool,
        translate_native_tables: bool,
        translate_ole_tables: bool,
    ) -> None:
        """Process all relevant shapes on a single slide."""
        try:
            for shape in slide.shapes:
                self._process_shape(
                    shape=shape,
                    slide_number=slide_number,
                    slide_width_emu=slide_width_emu,
                    translate_fn=translate_fn,
                    mirror_positions=mirror_positions,
                    translate_charts=translate_charts,
                    translate_native_tables=translate_native_tables,
                    translate_ole_tables=translate_ole_tables,
                )
        except Exception as exc:
            logger.error(
                '_process_slide: error on slide %d: %s', slide_number, exc, exc_info=True
            )
            self._report.error(f'Slide {slide_number}: {exc}')

    def _process_shape(
        self,
        shape: Any,
        slide_number: int,
        slide_width_emu: int,
        translate_fn: Callable[[str], str],
        mirror_positions: bool,
        translate_charts: bool,
        translate_native_tables: bool,
        translate_ole_tables: bool,
    ) -> None:
        """Process a single shape for embedded Excel content."""
        shape_name = getattr(shape, 'name', '<unknown>')
        try:
            kind = classify_shape(shape)

            if kind == 'ole_excel_table' and translate_ole_tables:
                self._report.total_ole_tables_found += 1
                logger.info(
                    'Slide %d: processing OLE Excel table "%s"',
                    slide_number, shape_name
                )
                # Build info for audit logging
                elem = shape.element
                is_embedded = _ole_is_embedded(elem)
                has_preview = any(True for _ in elem.iter(f'{{{P_NS}}}blipFill'))
                ole_info = EmbeddedExcelInfo(
                    slide_number=slide_number,
                    shape_id=getattr(shape, 'shape_id', 0),
                    shape_name=shape_name,
                    object_type='ole_excel_table',
                    prog_id=_get_ole_prog_id(elem),
                    is_embedded=is_embedded,
                    has_preview_image=has_preview,
                )
                if not is_embedded:
                    self._report.log_audit(
                        'OLE_LINKED_SKIP',
                        f'Slide {slide_number}: OLE table "{shape_name}" is linked; skipping translation'
                    )
                    if mirror_positions:
                        self.mirror_shape_position(shape, slide_width_emu)
                else:
                    n = self.translate_ole_excel_table(shape, translate_fn, ole_info)
                    if n >= 0:
                        self._report.total_ole_tables_translated += 1
                    if mirror_positions:
                        self.mirror_shape_position(shape, slide_width_emu)

            elif kind == 'native_chart' and translate_charts:
                self._report.total_charts_found += 1
                logger.info(
                    'Slide %d: processing native chart "%s"',
                    slide_number, shape_name
                )
                n = self.translate_chart_labels(shape, translate_fn)
                self.apply_rtl_to_chart(shape)
                if mirror_positions:
                    self.mirror_shape_position(shape, slide_width_emu)
                self._report.total_charts_translated += 1

            elif kind == 'ole_excel_chart':
                # Legacy OLE charts — limited access without OLE server
                self._report.log_audit(
                    'OLE_CHART_SKIP',
                    f'Slide {slide_number}: legacy OLE chart "{shape_name}" — '
                    f'text translation not available; applying position mirror only'
                )
                if mirror_positions:
                    self.mirror_shape_position(shape, slide_width_emu)

            elif kind == 'native_table' and translate_native_tables:
                logger.info(
                    'Slide %d: processing native table "%s"',
                    slide_number, shape_name
                )
                self.process_native_table(
                    shape, translate_fn, slide_width_emu,
                    mirror_position=mirror_positions
                )

            elif kind == 'picture' and mirror_positions:
                # Reposition static image shapes
                self.mirror_shape_position(shape, slide_width_emu)
                self._report.log_audit(
                    'SKIP_PICTURE',
                    f'Slide {slide_number}: picture "{shape_name}" repositioned only '
                    f'(no text translation on static images)'
                )

        except Exception as exc:
            msg = (
                f'_process_shape: unexpected error on slide {slide_number}, '
                f'shape "{shape_name}": {exc}'
            )
            logger.error(msg, exc_info=True)
            self._report.error(msg)

    # ─────────────────────────────────────────────────────────────────────────
    # ZIP-based fallback for embedded part replacement
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def replace_embedded_part_via_zip(
        pptx_path: str,
        part_path: str,
        new_bytes: bytes,
        output_path: str,
    ) -> bool:
        """Replace an embedded part in a PPTX by treating it as a ZIP archive.

        This is the fallback when python-pptx private API write-back fails.
        The PPTX file is unpacked to a temporary directory, the target part
        is overwritten, and the archive is rebuilt.

        CAUTION: This bypasses python-pptx's in-memory representation.
        Only use after saving the presentation first, or on a file path directly.

        Args:
            pptx_path: Path to the source PPTX file.
            part_path: Internal ZIP path to replace (e.g.,
                'ppt/embeddings/Microsoft_Excel_Worksheet1.xlsx').
            new_bytes: New bytes to write to the part.
            output_path: Path to write the modified PPTX.

        Returns:
            True on success, False on failure.
        """
        try:
            tmp_dir = tempfile.mkdtemp(prefix='slideshift_zip_')
            try:
                with zipfile.ZipFile(pptx_path, 'r') as zin:
                    zin.extractall(tmp_dir)

                # Overwrite the target part
                target = os.path.join(tmp_dir, part_path.replace('/', os.sep))
                target_dir = os.path.dirname(target)
                os.makedirs(target_dir, exist_ok=True)
                with open(target, 'wb') as f:
                    f.write(new_bytes)

                # Repack to output_path
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for root, dirs, files in os.walk(tmp_dir):
                        for filename in files:
                            filepath = os.path.join(root, filename)
                            arcname = os.path.relpath(filepath, tmp_dir)
                            zout.write(filepath, arcname)

                logger.info(
                    'replace_embedded_part_via_zip: wrote %s → %s',
                    part_path, output_path
                )
                return True

            finally:
                shutil.rmtree(tmp_dir, ignore_errors=True)

        except Exception as exc:
            logger.error(
                'replace_embedded_part_via_zip: failed for %s: %s',
                part_path, exc, exc_info=True
            )
            return False

    # ─────────────────────────────────────────────────────────────────────────
    # Public report accessor
    # ─────────────────────────────────────────────────────────────────────────

    @property
    def report(self) -> EmbeddedExcelHandlerReport:
        """Access the current processing report."""
        return self._report
